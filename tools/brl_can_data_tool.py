#!/usr/bin/env python3
"""
BRL Can Data Tool
=================

Vereintes Werkzeug fuer CAN-Checked Sensorkonfigurationen:

  - .TRX (verschluesselt, AES-256-ECB) einlesen
  - .TRI (Klartext) einlesen
  - .brl Fahrzeugprofile lesen + erstellen
  - Sensor-Vorschau mit Filter + Detail-Panel
  - Excel-Export (ein Blatt pro Datei + Uebersicht)
  - BRL-Profil aus TRX/TRI erzeugen (Fahrzeugdaten-Formular)

Ersetzt die Einzeltools trx_to_excel_gui.py und brl_format_gui.py.

Voraussetzungen:
  pip install pycryptodome openpyxl

Start:
  python3 brl_can_data_tool.py
"""

import json
import os
import struct
import subprocess
import sys
import threading
import zlib
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

APP_VERSION = "1.0.0"
APP_TITLE   = "BRL Can Data Tool"

# -- Script-Verzeichnis (PyInstaller-freundlich) ----------------------------
SCRIPT_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))

# -- AES-Schluessel ----------------------------------------------------------
# TRX: AES-256-ECB, aus CAN-Checked MFD15 Firmware
TRX_KEY = bytes.fromhex(
    "3373357638792f423f4528482b4d6251655468576d5a7134743777397a244326"
)
# BRL: AES-256-CBC, muss identisch in der Firmware sein
BRL_KEY = bytes.fromhex(
    "4d358275a7ea02f81df7c5689c073f8edbd464066f563717e6cbaf60addb710a"
)
BRL_MAGIC       = b"BRL\x01"
BRL_VERSION     = 1
BRL_HEADER_SIZE = 32
TRX_RECORD_SIZE      = 192
TRX_SENSOR_TEXT_SIZE = 96

# -- Kanonische Sensor-Feldreihenfolge (26 Felder) --------------------------
# Interne Schluessel sind die kurzen englischen Namen (wie im BRL-JSON).
# Fuer UI/Excel werden sie via FIELD_LABELS_DE uebersetzt.
FIELD_ORDER = [
    "proto", "can_id", "fmt", "start", "len",
    "unsigned", "shift", "mask", "decimals", "name",
    "scale", "offset", "mapper_type", "mapper1", "mapper2",
    "mapper3", "mapper4", "ain", "min", "max",
    "ref_sensor", "ref_val", "unused1", "popup", "unused2", "type",
]
FIELD_LABELS_DE = {
    "proto":        "Protokoll",
    "can_id":       "CAN-ID",
    "fmt":          "Format",
    "start":        "Startbyte",
    "len":          "Laenge",
    "unsigned":     "Unsigned",
    "shift":        "Shift",
    "mask":         "CAN-Maske",
    "decimals":     "Dezimalstellen",
    "name":         "Sensorname",
    "scale":        "Skalierung",
    "offset":       "Offset",
    "mapper_type":  "Mapper-Typ",
    "mapper1":      "Mapper-Info1",
    "mapper2":      "Mapper-Info2",
    "mapper3":      "Mapper-Info3",
    "mapper4":      "Mapper-Info4",
    "ain":          "AIN aktiv",
    "min":          "Min-Warnung",
    "max":          "Max-Warnung",
    "ref_sensor":   "Ref-Sensor",
    "ref_val":      "Ref-Wert",
    "unused1":      "(ungenutzt1)",
    "popup":        "Popup",
    "unused2":      "(ungenutzt2)",
    "type":         "Sensor-Typ",
}
PROTOCOL_LABELS = {
    "0":    "PT-CAN Broadcast",
    "0000": "PT-CAN Broadcast",
    "1":    "BMW UDS Diagnose",
    "0001": "BMW UDS Diagnose",
    "7DF":  "OBD2 Broadcast",
    "7df":  "OBD2 Broadcast",
    "FFF":  "MFD Analogeingang / Intern",
    "fff":  "MFD Analogeingang / Intern",
}
SENSOR_TYPE_LABELS = {
    "0": "-",
    "1": "Druck",
    "2": "Temperatur",
    "3": "Geschwindigkeit",
    "4": "Lambda",
}

# -- Farbschema --------------------------------------------------------------
COL_BG      = "#000000"
COL_PANEL   = "#1A1A1A"
COL_PANEL2  = "#111111"
COL_BORDER  = "#2C2C2E"
COL_BLUE    = "#1C69D4"
COL_BLUE_DK = "#0D2B5A"
COL_WHITE   = "#FFFFFF"
COL_GRAY    = "#AEAEB2"
COL_DARK    = "#636366"
COL_GREEN   = "#30D158"
COL_AMBER   = "#FF9F0A"
COL_RED     = "#FF453A"


# ===========================================================================
# Format-Parser
# ===========================================================================

def _pkcs7_unpad(data: bytes) -> bytes:
    if not data:
        raise ValueError("leerer Block")
    pad = data[-1]
    if pad < 1 or pad > 16 or data[-pad:] != bytes([pad]) * pad:
        raise ValueError(f"ungueltiges PKCS7-Padding ({pad})")
    return data[:-pad]


def _pkcs7_pad(data: bytes) -> bytes:
    pad = 16 - (len(data) % 16)
    return data + bytes([pad] * pad)


def _split_header(data: bytes) -> tuple[bytes, bytes]:
    """TRX/TRI-Header (erste Zeile) von Body trennen."""
    hdr_end = data.find(b"\r\n")
    if hdr_end != -1:
        return data[: hdr_end + 2], data[hdr_end + 2 :]
    hdr_end = data.find(b"\n")
    if hdr_end != -1:
        return data[: hdr_end + 1], data[hdr_end + 1 :]
    raise ValueError("kein Zeilenende im Header")


def _sensor_dict_from_fields(parts: list[str], slot: int) -> dict:
    """Fuellt ein Sensor-Dict aus 26 Semikolon-getrennten Feldern."""
    while len(parts) < 26:
        parts.append("")
    s = {FIELD_ORDER[j]: parts[j] for j in range(26)}
    s["_slot"]  = slot
    s["_empty"] = False
    s["_leer"]  = s.get("name", "").lower() == "empty"
    # Klartext-Helper fuer UI (nicht in BRL geschrieben)
    s["_proto_text"] = PROTOCOL_LABELS.get(s.get("proto", ""), s.get("proto", ""))
    s["_type_text"]  = SENSOR_TYPE_LABELS.get(s.get("type", ""),  s.get("type", ""))
    return s


def _parse_trx_records(plaintext: bytes) -> list[dict]:
    """192-Byte-Records in Sensor-Dicts wandeln (gilt TRX + TRI)."""
    sensors: list[dict] = []
    num_records = len(plaintext) // TRX_RECORD_SIZE
    for i in range(num_records):
        block = plaintext[i * TRX_RECORD_SIZE :
                          i * TRX_RECORD_SIZE + TRX_SENSOR_TEXT_SIZE]
        null = block.find(b"\x00")
        text = block[: null if null != -1 else TRX_SENSOR_TEXT_SIZE]
        text = text.decode("latin-1", errors="replace").rstrip(";")
        if not text or ";" not in text:
            sensors.append({"_slot": i, "_empty": True})
            continue
        sensors.append(_sensor_dict_from_fields(text.split(";"), i))
    return sensors


def _parse_tri_text(body: bytes) -> list[dict]:
    """TRI ist Klartext: zeilenweise, eine Zeile pro Sensor."""
    sensors: list[dict] = []
    lines = body.replace(b"\r\n", b"\n").split(b"\n")
    for i, line in enumerate(lines):
        text = line.decode("latin-1", errors="replace").rstrip(";").strip()
        if not text or ";" not in text:
            continue
        sensors.append(_sensor_dict_from_fields(text.split(";"), i))
    return sensors


def read_trx(path: Path) -> tuple[str, list[dict]]:
    from Crypto.Cipher import AES
    data = path.read_bytes()
    header, body = _split_header(data)
    if not body:
        raise ValueError("Body leer")
    if len(body) % 16 != 0:
        raise ValueError(f"Body-Laenge {len(body)} nicht durch 16 teilbar")
    plaintext = AES.new(TRX_KEY, AES.MODE_ECB).decrypt(body)
    return (header.decode("latin-1", errors="replace").strip(),
            _parse_trx_records(plaintext))


def read_tri(path: Path) -> tuple[str, list[dict]]:
    data = path.read_bytes()
    # Tolerant: wenn ein Header da ist, abtrennen; sonst direkt Body parsen.
    try:
        header, body = _split_header(data)
        header_text = header.decode("latin-1", errors="replace").strip()
    except ValueError:
        header_text = ""
        body        = data
    # TRI kann sowohl als Zeilen-Text (mit \n) als auch als 192-Byte-Records
    # (decrypt_trx schrieb frueher so) vorliegen. Zeilen-Format bevorzugen
    # wenn Body nicht durch 192 teilbar ist.
    if len(body) >= TRX_RECORD_SIZE and len(body) % TRX_RECORD_SIZE == 0:
        sensors = _parse_trx_records(body)
    else:
        sensors = _parse_tri_text(body)
    return header_text, sensors


def read_brl(path: Path) -> tuple[dict, list[dict]]:
    from Crypto.Cipher import AES
    data = path.read_bytes()
    if len(data) < BRL_HEADER_SIZE or data[:4] != BRL_MAGIC:
        raise ValueError("kein gueltiger BRL-Header")
    iv           = data[8:24]
    payload_size = struct.unpack_from("<I", data, 24)[0]
    expected_crc = struct.unpack_from("<I", data, 28)[0]
    encrypted    = data[BRL_HEADER_SIZE : BRL_HEADER_SIZE + payload_size]
    if len(encrypted) != payload_size:
        raise ValueError("unvollstaendige Payload")
    actual_crc = zlib.crc32(encrypted) & 0xFFFFFFFF
    if actual_crc != expected_crc:
        raise ValueError(
            f"CRC32-Fehler: erwartet {expected_crc:08X}, bekommen {actual_crc:08X}"
        )
    payload = json.loads(
        _pkcs7_unpad(AES.new(BRL_KEY, AES.MODE_CBC, iv).decrypt(encrypted)).decode("utf-8")
    )
    vehicle = {k: payload.get(k, "") for k in
               ("make", "model", "engine", "name", "year_from",
                "year_to", "can", "bitrate")}
    sensors_raw = payload.get("sensors", [])
    # In _sensor_dict_from_fields Form bringen (Slot + _* helpers)
    sensors: list[dict] = []
    for i, raw in enumerate(sensors_raw):
        s = {k: str(raw.get(k, "")) for k in FIELD_ORDER}
        s["_slot"]       = int(raw.get("slot", i))
        s["_empty"]      = False
        s["_leer"]       = s.get("name", "").lower() == "empty"
        s["_proto_text"] = PROTOCOL_LABELS.get(s["proto"], s["proto"])
        s["_type_text"]  = SENSOR_TYPE_LABELS.get(s["type"],  s["type"])
        sensors.append(s)
    return vehicle, sensors


def create_brl(vehicle: dict, sensors: list[dict]) -> bytes:
    """Schreibt eine .brl Datei. Sensoren werden auf BRL-JSON-Shape reduziert."""
    from Crypto.Cipher import AES
    clean_sensors: list[dict] = []
    for s in sensors:
        if s.get("_empty") or s.get("_leer"):
            continue
        entry = {k: s.get(k, "") for k in FIELD_ORDER}
        entry["slot"] = s.get("_slot", 0)
        clean_sensors.append(entry)
    payload = {
        "v":         BRL_VERSION,
        "make":      vehicle.get("make", ""),
        "model":     vehicle.get("model", ""),
        "engine":    vehicle.get("engine", ""),
        "name":      vehicle.get("name", ""),
        "year_from": int(vehicle.get("year_from") or 0),
        "year_to":   int(vehicle.get("year_to") or 9999),
        "can":       vehicle.get("can", "PT-CAN"),
        "bitrate":   int(vehicle.get("bitrate") or 500),
        "sensors":   clean_sensors,
    }
    json_bytes = json.dumps(payload, ensure_ascii=False,
                            separators=(",", ":")).encode("utf-8")
    iv = os.urandom(16)
    encrypted = AES.new(BRL_KEY, AES.MODE_CBC, iv).encrypt(_pkcs7_pad(json_bytes))
    crc = zlib.crc32(encrypted) & 0xFFFFFFFF
    header = (
        BRL_MAGIC
        + bytes([BRL_VERSION, 0, 0, 0])
        + iv
        + struct.pack("<I", len(encrypted))
        + struct.pack("<I", crc)
    )
    return header + encrypted


# ===========================================================================
# Vereinheitlichtes Lade-Objekt
# ===========================================================================

class LoadedFile:
    """Repraesentiert eine geladene Sensor-Quelle (.TRX/.TRI/.brl)."""

    def __init__(self, path: Path):
        self.path: Path                 = path
        self.kind: str                  = ""          # "TRX" / "TRI" / "BRL"
        self.header: str                = ""          # nur TRX/TRI
        self.vehicle: dict | None       = None        # nur BRL
        self.sensors: list[dict]        = []
        self.error: str | None          = None
        self._detect_and_load()

    def _detect_and_load(self) -> None:
        ext = self.path.suffix.lower()
        try:
            if ext == ".brl":
                self.kind = "BRL"
                self.vehicle, self.sensors = read_brl(self.path)
                return
            if ext == ".trx":
                self.kind = "TRX"
                self.header, self.sensors = read_trx(self.path)
                return
            if ext == ".tri":
                self.kind = "TRI"
                self.header, self.sensors = read_tri(self.path)
                return
            # Unbekannte Endung: Magic schnueffeln
            head = self.path.read_bytes()[:4]
            if head == BRL_MAGIC:
                self.kind = "BRL"
                self.vehicle, self.sensors = read_brl(self.path)
            else:
                # TRX oder TRI: zuerst TRX probieren, bei Fehler Fallback TRI
                try:
                    self.header, self.sensors = read_trx(self.path)
                    self.kind = "TRX"
                except Exception:
                    self.header, self.sensors = read_tri(self.path)
                    self.kind = "TRI"
        except Exception as e:
            self.error = str(e)

    def active_sensor_count(self) -> int:
        return sum(1 for s in self.sensors
                   if not s.get("_empty") and not s.get("_leer"))


# ===========================================================================
# Excel-Export
# ===========================================================================

EXCEL_COLS = (
    ["Slot", FIELD_LABELS_DE["name"], FIELD_LABELS_DE["proto"],
     "Protokoll (Klartext)", FIELD_LABELS_DE["can_id"]]
    + [FIELD_LABELS_DE[k] for k in FIELD_ORDER[2:]]  # alle Felder ab fmt
    + ["Sensor-Typ (Klartext)"]
)


def _sensor_row_values(loaded: LoadedFile, s: dict) -> list:
    row = [
        s.get("_slot", 0) + 1,
        s.get("name", ""),
        s.get("proto", ""),
        s.get("_proto_text", ""),
        s.get("can_id", ""),
    ]
    for k in FIELD_ORDER[2:]:
        row.append(s.get(k, ""))
    row.append(s.get("_type_text", ""))
    return row


def export_excel(loaded_files: list[LoadedFile], output_path: Path) -> int:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    META_FILL = PatternFill("solid", fgColor="D6E4F0")
    EVEN_FILL = PatternFill("solid", fgColor="F2F7FB")
    THIN      = Side(style="thin", color="CCCCCC")
    TB        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_hdr(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = TB
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
        ws.row_dimensions[row].height = 30

    def style_data(ws, row, ncols, even, grey):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            if grey:
                cell.font = Font(color="AAAAAA", italic=True)
            if even:
                cell.fill = EVEN_FILL
            cell.border = TB
            cell.alignment = Alignment(vertical="center")

    overview_cols = ["Datei"] + EXCEL_COLS
    ws_ov = wb.create_sheet("Uebersicht")
    ws_ov.freeze_panes = "A3"
    ws_ov["A1"] = "Uebersicht aller aktiven Sensoren"
    ws_ov["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws_ov.merge_cells(f"A1:{get_column_letter(len(overview_cols))}1")
    for ci, col in enumerate(overview_cols, 1):
        ws_ov.cell(row=2, column=ci, value=col)
    style_hdr(ws_ov, 2, len(overview_cols))
    ov_row = 3

    for lf in loaded_files:
        if lf.error:
            continue
        sheet = lf.path.stem[:31]
        used = [s.title for s in wb.worksheets]
        if sheet in used:
            sheet = sheet[:27] + f"_{len(used)}"
        ws = wb.create_sheet(sheet)
        ws.freeze_panes = "A4"

        meta = f"Quelle: {lf.path.name} ({lf.kind})"
        if lf.kind == "BRL" and lf.vehicle:
            meta += (f"  -  {lf.vehicle.get('make','')} "
                     f"{lf.vehicle.get('model','')}  "
                     f"Motor: {lf.vehicle.get('engine','')}")
        ws["A1"] = meta
        ws["A1"].font = Font(bold=True, size=10)
        ws["A1"].fill = META_FILL
        ws.merge_cells(f"A1:{get_column_letter(len(EXCEL_COLS))}1")

        second = lf.header if lf.kind in ("TRX", "TRI") else (
            json.dumps(lf.vehicle, ensure_ascii=False) if lf.vehicle else ""
        )
        ws["A2"] = second
        ws["A2"].font = Font(italic=True, size=9, color="555555")
        ws.merge_cells(f"A2:{get_column_letter(len(EXCEL_COLS))}2")

        for ci, col in enumerate(EXCEL_COLS, 1):
            ws.cell(row=3, column=ci, value=col)
        style_hdr(ws, 3, len(EXCEL_COLS))

        row = 4
        for s in lf.sensors:
            if s.get("_empty"):
                continue
            is_leer = s.get("_leer", False)
            vals = _sensor_row_values(lf, s)
            for ci, v in enumerate(vals, 1):
                ws.cell(row=row, column=ci, value=v)
            style_data(ws, row, len(EXCEL_COLS), row % 2 == 0, is_leer)
            if not is_leer:
                ov_vals = [lf.path.name] + vals
                for ci, v in enumerate(ov_vals, 1):
                    ws_ov.cell(row=ov_row, column=ci, value=v)
                style_data(ws_ov, ov_row, len(overview_cols),
                           ov_row % 2 == 0, False)
                ov_row += 1
            row += 1

        widths = {1: 6, 2: 20, 3: 12, 4: 22, 5: 14}
        for ci in range(1, len(EXCEL_COLS) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(ci, 14)

    for ci, w in enumerate([20, 6, 20, 12, 22, 14], 1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(7, len(overview_cols) + 1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = 14

    wb.move_sheet("Uebersicht", offset=-len(wb.sheetnames) + 1)
    wb.save(output_path)
    return ov_row - 3


# ===========================================================================
# GUI
# ===========================================================================

class App(tk.Tk):

    # -- Init ----------------------------------------------------------------

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.configure(bg=COL_BG)
        self.minsize(1080, 680)

        self._files: list[LoadedFile] = []

        self._build_style()
        self._build_header()
        self._build_main()
        self._check_deps()
        self._refresh_files()

    def _build_style(self):
        st = ttk.Style(self)
        try:
            st.theme_use("clam")
        except Exception:
            pass
        st.configure("BMW.TNotebook",
                     background=COL_BG, borderwidth=0, tabmargins=(4, 4, 0, 0))
        st.configure("BMW.TNotebook.Tab",
                     background=COL_PANEL, foreground=COL_GRAY,
                     padding=[14, 7], borderwidth=0,
                     font=("Helvetica", 10))
        st.map("BMW.TNotebook.Tab",
               background=[("selected", COL_BG), ("active", COL_BORDER)],
               foreground=[("selected", COL_WHITE)])
        st.configure("BMW.Treeview",
                     background=COL_PANEL2, fieldbackground=COL_PANEL2,
                     foreground=COL_GRAY, bordercolor=COL_BORDER,
                     rowheight=22, font=("Helvetica", 10))
        st.configure("BMW.Treeview.Heading",
                     background=COL_PANEL, foreground=COL_WHITE,
                     relief="flat", font=("Helvetica", 10, "bold"))
        st.map("BMW.Treeview.Heading",
               background=[("active", COL_BORDER)])
        st.map("BMW.Treeview",
               background=[("selected", COL_BLUE)],
               foreground=[("selected", COL_WHITE)])
        st.configure("BMW.Vertical.TScrollbar",
                     background=COL_PANEL, troughcolor=COL_BG,
                     bordercolor=COL_BORDER, arrowcolor=COL_DARK)

    def _build_header(self):
        hdr = tk.Frame(self, bg=COL_BG)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=COL_BLUE, width=6).pack(side="left", fill="y")
        logo_f = tk.Frame(hdr, bg=COL_BG, padx=14, pady=10)
        logo_f.pack(side="left")
        self._draw_logo(logo_f)
        title_f = tk.Frame(hdr, bg=COL_BG, padx=8, pady=10)
        title_f.pack(side="left", fill="both", expand=True)
        tk.Label(title_f, text="BRL CAN DATA TOOL",
                 bg=COL_BG, fg=COL_WHITE,
                 font=("Helvetica", 18, "bold")).pack(anchor="w")
        tk.Label(title_f,
                 text="TRX / TRI / BRL einlesen  -  Sensor-Vorschau  -  "
                      "Excel-Export  -  Fahrzeugprofile erstellen",
                 bg=COL_BG, fg=COL_GRAY, font=("Helvetica", 9)).pack(anchor="w")
        tk.Label(hdr, text=f"v{APP_VERSION}",
                 bg=COL_BG, fg=COL_BLUE,
                 font=("Helvetica", 11, "bold"),
                 padx=18).pack(side="right", fill="y")
        tk.Frame(self, bg=COL_BORDER, height=1).pack(fill="x")

    def _draw_logo(self, parent):
        logo_path = SCRIPT_DIR / "logo.png"
        if logo_path.exists():
            try:
                img = tk.PhotoImage(file=str(logo_path))
                factor = max(1, max(img.width(), img.height()) // 64)
                if factor > 1:
                    img = img.subsample(factor, factor)
                self._logo_img = img
                tk.Label(parent, image=img, bg=COL_BG).pack()
                return
            except Exception:
                pass
        size, m = 58, 2
        r  = size // 2
        ri = r - 6
        cv = tk.Canvas(parent, width=size, height=size,
                       bg=COL_BG, highlightthickness=0)
        cv.pack()
        cv.create_oval(m, m, size - m, size - m,
                       fill="#0D1B2A", outline=COL_BLUE, width=2)
        for start, col in ((90, COL_BLUE), (0, COL_WHITE),
                           (270, COL_BLUE), (180, COL_WHITE)):
            cv.create_arc(r - ri, r - ri, r + ri, r + ri,
                          start=start, extent=90,
                          fill=col, outline="", style="pieslice")
        cv.create_line(r, r - ri, r, r + ri, fill="#0D1B2A", width=2)
        cv.create_line(r - ri, r, r + ri, r, fill="#0D1B2A", width=2)

    def _build_main(self):
        root = tk.Frame(self, bg=COL_BG)
        root.pack(fill="both", expand=True, padx=12, pady=10)
        # Zweispaltiges Layout: links Dateien, rechts Tabs
        paned = tk.PanedWindow(root, orient="horizontal",
                               bg=COL_BG, bd=0, sashwidth=6,
                               sashrelief="flat")
        paned.pack(fill="both", expand=True)

        # -- linke Spalte: Datei-Liste + Buttons ---------------------------
        left = tk.Frame(paned, bg=COL_PANEL,
                        highlightbackground=COL_BORDER,
                        highlightthickness=1)
        tk.Label(left, text="  DATEIEN",
                 bg=COL_PANEL, fg=COL_BLUE,
                 font=("Helvetica", 9, "bold")).pack(anchor="w",
                                                      padx=10, pady=(10, 4))
        btns = tk.Frame(left, bg=COL_PANEL)
        btns.pack(fill="x", padx=10, pady=(0, 6))
        self._mk_btn(btns, "Datei hinzufuegen",
                     self._add_files, COL_BLUE).pack(side="left", padx=(0, 4))
        self._mk_btn(btns, "Ordner",
                     self._add_folder, COL_BLUE_DK).pack(side="left", padx=(0, 4))
        self._mk_btn(btns, "Liste leeren",
                     self._clear_files, "#3A1010",
                     hover="#5A1515").pack(side="right")

        lf_frame = tk.Frame(left, bg=COL_PANEL)
        lf_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        self.tree_files = ttk.Treeview(lf_frame, style="BMW.Treeview",
                                       columns=("kind", "sensors"),
                                       show="tree headings")
        self.tree_files.heading("#0",      text="Datei")
        self.tree_files.heading("kind",    text="Typ")
        self.tree_files.heading("sensors", text="Sensoren")
        self.tree_files.column("#0", width=220, anchor="w")
        self.tree_files.column("kind", width=52, anchor="center")
        self.tree_files.column("sensors", width=80, anchor="e")
        sb = ttk.Scrollbar(lf_frame, orient="vertical",
                           command=self.tree_files.yview,
                           style="BMW.Vertical.TScrollbar")
        self.tree_files.configure(yscrollcommand=sb.set)
        self.tree_files.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tree_files.bind("<<TreeviewSelect>>", self._on_select_file)

        self.lbl_count = tk.Label(left, text="0 Dateien",
                                  bg=COL_PANEL, fg=COL_DARK,
                                  font=("Helvetica", 9))
        self.lbl_count.pack(anchor="w", padx=12, pady=(0, 8))

        # -- rechte Spalte: Tabs ------------------------------------------
        right = tk.Frame(paned, bg=COL_BG)
        self.tabs = ttk.Notebook(right, style="BMW.TNotebook")
        self.tabs.pack(fill="both", expand=True)

        self.tab_preview = tk.Frame(self.tabs, bg=COL_BG)
        self.tab_excel   = tk.Frame(self.tabs, bg=COL_BG)
        self.tab_brl     = tk.Frame(self.tabs, bg=COL_BG)
        self.tabs.add(self.tab_preview, text="  Sensor-Vorschau  ")
        self.tabs.add(self.tab_excel,   text="  Excel Export  ")
        self.tabs.add(self.tab_brl,     text="  BRL Profil  ")

        self._build_preview_tab()
        self._build_excel_tab()
        self._build_brl_tab()

        paned.add(left,  minsize=280)
        paned.add(right, minsize=640)

    # -- Preview Tab ---------------------------------------------------------

    def _build_preview_tab(self):
        p = self.tab_preview

        toolbar = tk.Frame(p, bg=COL_BG)
        toolbar.pack(fill="x", padx=4, pady=(6, 4))
        tk.Label(toolbar, text="Filter:",
                 bg=COL_BG, fg=COL_GRAY,
                 font=("Helvetica", 10)).pack(side="left")
        self.var_filter = tk.StringVar()
        self.var_filter.trace_add("write", lambda *_: self._refresh_preview())
        tk.Entry(toolbar, textvariable=self.var_filter,
                 bg=COL_PANEL, fg=COL_WHITE, insertbackground=COL_WHITE,
                 relief="flat", font=("Helvetica", 10),
                 highlightbackground=COL_BORDER, highlightthickness=1
                 ).pack(side="left", fill="x", expand=True, padx=8)
        self.var_hide_empty = tk.BooleanVar(value=True)
        tk.Checkbutton(toolbar, text="Leere Sensoren ausblenden",
                       variable=self.var_hide_empty,
                       command=self._refresh_preview,
                       bg=COL_BG, fg=COL_GRAY,
                       selectcolor=COL_PANEL,
                       activebackground=COL_BG, activeforeground=COL_WHITE,
                       font=("Helvetica", 9)).pack(side="left", padx=(12, 0))

        # Preview-Treeview (alle Sensoren der gerade ausgewaehlten Datei)
        pv_frame = tk.Frame(p, bg=COL_BG)
        pv_frame.pack(fill="both", expand=True, padx=4, pady=4)

        cols = ["slot", "name", "proto_text", "can_id", "len", "scale",
                "offset", "min", "max", "type_text"]
        headings = {
            "slot":       "#",
            "name":       FIELD_LABELS_DE["name"],
            "proto_text": "Protokoll",
            "can_id":     FIELD_LABELS_DE["can_id"],
            "len":        FIELD_LABELS_DE["len"],
            "scale":      FIELD_LABELS_DE["scale"],
            "offset":     FIELD_LABELS_DE["offset"],
            "min":        FIELD_LABELS_DE["min"],
            "max":        FIELD_LABELS_DE["max"],
            "type_text":  "Typ",
        }
        widths = {"slot": 40, "name": 170, "proto_text": 160,
                  "can_id": 90, "len": 60, "scale": 90, "offset": 80,
                  "min": 70, "max": 70, "type_text": 110}
        self.tree_sens = ttk.Treeview(pv_frame, style="BMW.Treeview",
                                      columns=cols, show="headings")
        for c in cols:
            self.tree_sens.heading(c, text=headings[c])
            self.tree_sens.column(c, width=widths[c],
                                  anchor="e" if c in ("slot", "len", "scale",
                                                       "offset", "min", "max")
                                          else "w")
        sb = ttk.Scrollbar(pv_frame, orient="vertical",
                           command=self.tree_sens.yview,
                           style="BMW.Vertical.TScrollbar")
        self.tree_sens.configure(yscrollcommand=sb.set)
        self.tree_sens.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tree_sens.bind("<<TreeviewSelect>>", self._on_select_sensor)

        # Detail-Panel unter der Preview
        detail = tk.LabelFrame(p, text="  SENSOR-DETAILS  ",
                               bg=COL_PANEL, fg=COL_BLUE,
                               font=("Helvetica", 9, "bold"),
                               bd=1, relief="flat",
                               highlightbackground=COL_BORDER,
                               highlightthickness=1)
        detail.pack(fill="x", padx=4, pady=(4, 6))
        self.txt_detail = tk.Text(detail, height=8, bg=COL_PANEL2,
                                  fg=COL_GRAY, font=("Consolas", 9),
                                  relief="flat", bd=0, padx=10, pady=8,
                                  state="disabled", wrap="word")
        self.txt_detail.pack(fill="x", padx=8, pady=6)

        # Status fuer BRL-Fahrzeuginfo
        self.lbl_vehicle = tk.Label(p, text="",
                                    bg=COL_BG, fg=COL_AMBER,
                                    font=("Helvetica", 9),
                                    anchor="w", justify="left")
        self.lbl_vehicle.pack(fill="x", padx=4, pady=(0, 6))

    # -- Excel Tab -----------------------------------------------------------

    def _build_excel_tab(self):
        p = self.tab_excel
        tk.Label(p, text="Exportiert alle geladenen Dateien als Excel-Tabelle:\n"
                         "- ein Tabellenblatt je Datei (alle Sensoren)\n"
                         "- ein Uebersichtsblatt mit allen aktiven Sensoren\n"
                         "- .TRX, .TRI, und .brl werden gleich behandelt",
                 bg=COL_BG, fg=COL_GRAY, font=("Helvetica", 10),
                 anchor="w", justify="left").pack(fill="x", padx=14, pady=(14, 10))

        row = tk.Frame(p, bg=COL_BG)
        row.pack(fill="x", padx=14, pady=6)
        tk.Label(row, text="Ausgabe:", bg=COL_BG, fg=COL_GRAY,
                 font=("Helvetica", 10), width=10, anchor="w").pack(side="left")
        self.var_excel_out = tk.StringVar(
            value=str(Path.home() / "can_sensoren.xlsx"))
        tk.Entry(row, textvariable=self.var_excel_out,
                 bg=COL_PANEL, fg=COL_WHITE, insertbackground=COL_WHITE,
                 relief="flat", font=("Helvetica", 10),
                 highlightbackground=COL_BORDER,
                 highlightthickness=1).pack(side="left", fill="x",
                                            expand=True, padx=(0, 8), ipady=4)
        tk.Button(row, text="...", command=self._choose_excel_out,
                  bg=COL_PANEL, fg=COL_GRAY, activebackground=COL_BORDER,
                  relief="flat", padx=10, cursor="hand2").pack(side="right")

        self.btn_excel = self._mk_btn(p, "  EXCEL EXPORTIEREN",
                                      self._do_export_excel,
                                      COL_GREEN, hover="#1E8A3C",
                                      big=True)
        self.btn_excel.pack(anchor="w", padx=14, pady=10)
        self.lbl_excel = tk.Label(p, text="", bg=COL_BG, fg=COL_GRAY,
                                  font=("Helvetica", 10))
        self.lbl_excel.pack(fill="x", padx=14)
        self.pb_excel = ttk.Progressbar(p, mode="determinate")
        self.pb_excel.pack(fill="x", padx=14, pady=8)

    # -- BRL Tab -------------------------------------------------------------

    def _build_brl_tab(self):
        p = self.tab_brl
        tk.Label(p,
                 text="Waehle links eine .TRX/.TRI/.brl als Quelle und "
                      "fuelle die Fahrzeugdaten aus. Eine ausgewaehlte "
                      ".brl Datei befuellt das Formular automatisch.\n"
                      "Zum Kombinieren mehrerer Quellen (z.B. OBD2 11-Bit + "
                      "OBD 29-Bit Extended) auf 'Alle Dateien zusammenfuehren' "
                      "schalten.",
                 bg=COL_BG, fg=COL_GRAY,
                 font=("Helvetica", 10),
                 anchor="w", justify="left", wraplength=720
                 ).pack(fill="x", padx=14, pady=(14, 6))

        # -- Quelle (Einzel-Datei / Merge) -----------------------------------
        src_frame = tk.LabelFrame(p, text="  QUELLE  ",
                                  bg=COL_PANEL, fg=COL_BLUE,
                                  font=("Helvetica", 9, "bold"),
                                  bd=1, relief="flat",
                                  highlightbackground=COL_BORDER,
                                  highlightthickness=1)
        src_frame.pack(fill="x", padx=14, pady=(0, 8))
        self.var_src_mode = tk.StringVar(value="single")
        tk.Radiobutton(src_frame,
                       text="Nur ausgewaehlte Datei",
                       variable=self.var_src_mode, value="single",
                       command=self._update_brl_src_label,
                       bg=COL_PANEL, fg=COL_GRAY, selectcolor=COL_BG,
                       activebackground=COL_PANEL, activeforeground=COL_WHITE,
                       font=("Helvetica", 10)).pack(anchor="w", padx=12, pady=(6, 0))
        tk.Radiobutton(src_frame,
                       text="Alle geladenen Dateien zusammenfuehren "
                            "(OBD2 + OBD29 etc.)",
                       variable=self.var_src_mode, value="merge",
                       command=self._update_brl_src_label,
                       bg=COL_PANEL, fg=COL_GRAY, selectcolor=COL_BG,
                       activebackground=COL_PANEL, activeforeground=COL_WHITE,
                       font=("Helvetica", 10)).pack(anchor="w", padx=12, pady=(0, 6))
        self.lbl_brl_src = tk.Label(src_frame, text="",
                                    bg=COL_PANEL, fg=COL_AMBER,
                                    font=("Helvetica", 9), anchor="w")
        self.lbl_brl_src.pack(fill="x", padx=12, pady=(0, 8))

        form = tk.Frame(p, bg=COL_BG)
        form.pack(fill="x", padx=14, pady=(0, 6))

        def lbl(r, c, text):
            tk.Label(form, text=text, bg=COL_BG, fg=COL_GRAY,
                     font=("Helvetica", 10), anchor="e", width=14
                     ).grid(row=r, column=c * 2,
                            sticky="e", padx=(0, 6), pady=4)

        def entry(r, c, var, width=24):
            e = tk.Entry(form, textvariable=var,
                         bg=COL_PANEL, fg=COL_WHITE,
                         insertbackground=COL_WHITE,
                         relief="flat", font=("Helvetica", 10), width=width,
                         highlightbackground=COL_BORDER,
                         highlightthickness=1)
            e.grid(row=r, column=c * 2 + 1, sticky="w",
                   padx=(0, 16), pady=4, ipady=3)
            return e

        self.var_make    = tk.StringVar(value="BMW")
        self.var_model   = tk.StringVar()
        self.var_engine  = tk.StringVar()
        self.var_name    = tk.StringVar()
        self.var_year_f  = tk.StringVar(value="2000")
        self.var_year_t  = tk.StringVar(value="9999")
        self.var_can     = tk.StringVar(value="PT-CAN")
        self.var_bitrate = tk.StringVar(value="500")

        lbl(0, 0, "Hersteller:");     entry(0, 0, self.var_make)
        lbl(0, 1, "Modell:");         entry(0, 1, self.var_model)
        lbl(1, 0, "Motorcode *:");    entry(1, 0, self.var_engine)
        lbl(1, 1, "Bezeichnung *:");  entry(1, 1, self.var_name, width=28)
        lbl(2, 0, "Baujahr von:");    entry(2, 0, self.var_year_f, width=8)
        lbl(2, 1, "Baujahr bis:");    entry(2, 1, self.var_year_t, width=8)
        lbl(3, 0, "CAN-Bus:")
        ttk.OptionMenu(form, self.var_can, "PT-CAN",
                       "PT-CAN", "K-CAN", "F-CAN", "LIN", "MOST"
                       ).grid(row=3, column=1, sticky="w", pady=4)
        lbl(3, 1, "Bitrate (kbps):")
        ttk.OptionMenu(form, self.var_bitrate, "500",
                       "100", "250", "500", "1000"
                       ).grid(row=3, column=3, sticky="w", pady=4)

        row_out = tk.Frame(p, bg=COL_BG)
        row_out.pack(fill="x", padx=14, pady=(8, 4))
        tk.Label(row_out, text="Ausgabe:", bg=COL_BG, fg=COL_GRAY,
                 font=("Helvetica", 10), width=10, anchor="w").pack(side="left")
        self.var_brl_out = tk.StringVar()
        tk.Entry(row_out, textvariable=self.var_brl_out,
                 bg=COL_PANEL, fg=COL_WHITE, insertbackground=COL_WHITE,
                 relief="flat", font=("Helvetica", 10),
                 highlightbackground=COL_BORDER,
                 highlightthickness=1).pack(side="left", fill="x",
                                            expand=True, padx=(0, 8), ipady=4)
        tk.Button(row_out, text="...",
                  command=self._choose_brl_out,
                  bg=COL_PANEL, fg=COL_GRAY, activebackground=COL_BORDER,
                  relief="flat", padx=10, cursor="hand2").pack(side="right")

        btn_row = tk.Frame(p, bg=COL_BG)
        btn_row.pack(fill="x", padx=14, pady=8)
        self.btn_brl_create = self._mk_btn(
            btn_row, "  BRL ERSTELLEN",
            self._do_create_brl, COL_BLUE, hover=COL_BLUE_DK, big=True)
        self.btn_brl_create.pack(side="left")
        self.lbl_brl = tk.Label(p, text="", bg=COL_BG, fg=COL_GRAY,
                                font=("Helvetica", 10))
        self.lbl_brl.pack(fill="x", padx=14)

    # -- Button Helper -------------------------------------------------------

    @staticmethod
    def _mk_btn(parent, text, cmd, bg, hover=None, big=False):
        font = ("Helvetica", 11, "bold") if big else ("Helvetica", 10, "bold")
        padx = 20 if big else 12
        pady = 10 if big else 7
        return tk.Button(
            parent, text=text, command=cmd,
            bg=bg, fg=COL_WHITE,
            activebackground=hover or COL_BLUE_DK, activeforeground=COL_WHITE,
            font=font, relief="flat", padx=padx, pady=pady, cursor="hand2")

    # -- Deps Check ---------------------------------------------------------

    def _check_deps(self):
        missing = []
        try:
            from Crypto.Cipher import AES  # noqa
        except ImportError:
            missing.append("pycryptodome")
        try:
            import openpyxl  # noqa
        except ImportError:
            missing.append("openpyxl")
        if missing:
            messagebox.showerror(
                "Fehlende Pakete",
                "Bitte installieren:\n\n    "
                f"pip install {' '.join(missing)}\n\nDann neu starten.")
            self.destroy()

    # -- Datei-Aktionen -----------------------------------------------------

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Dateien auswaehlen",
            filetypes=[
                ("CAN-Dateien", "*.TRX *.trx *.TRI *.tri *.brl"),
                ("TRX (verschluesselt)", "*.TRX *.trx"),
                ("TRI (Klartext)",       "*.TRI *.tri"),
                ("BRL Fahrzeugprofil",   "*.brl"),
                ("Alle Dateien", "*.*"),
            ])
        for p in paths:
            self._append_file(Path(p))
        self._refresh_files()

    def _add_folder(self):
        d = filedialog.askdirectory(title="Ordner waehlen")
        if not d:
            return
        found = []
        for pat in ("*.TRX", "*.trx", "*.TRI", "*.tri", "*.brl"):
            found += sorted(Path(d).glob(pat))
        if not found:
            messagebox.showinfo("Keine Dateien",
                                f"Keine TRX/TRI/BRL in:\n{d}")
            return
        for p in found:
            self._append_file(p)
        self._refresh_files()

    def _append_file(self, p: Path):
        if any(lf.path == p for lf in self._files):
            return
        lf = LoadedFile(p)
        self._files.append(lf)

    def _clear_files(self):
        self._files.clear()
        self._refresh_files()

    def _refresh_files(self):
        self.tree_files.delete(*self.tree_files.get_children())
        for i, lf in enumerate(self._files):
            badge = lf.kind or "?"
            if lf.error:
                badge = "ERR"
                count = "-"
            else:
                count = str(lf.active_sensor_count())
            self.tree_files.insert(
                "", "end", iid=str(i),
                text=lf.path.name,
                values=(badge, count))
        total = sum(lf.active_sensor_count() for lf in self._files
                    if not lf.error)
        self.lbl_count.config(
            text=f"{len(self._files)} Datei(en), {total} Sensoren gesamt")
        self._refresh_preview()
        self._update_brl_src_label()

    # -- Selection handlers -------------------------------------------------

    def _selected_file(self) -> LoadedFile | None:
        sel = self.tree_files.selection()
        if not sel:
            return None
        idx = int(sel[0])
        if 0 <= idx < len(self._files):
            return self._files[idx]
        return None

    def _on_select_file(self, _evt=None):
        lf = self._selected_file()
        # BRL-Formular autofuellen
        if lf and lf.kind == "BRL" and lf.vehicle:
            v = lf.vehicle
            self.var_make.set(v.get("make", "BMW"))
            self.var_model.set(v.get("model", ""))
            self.var_engine.set(v.get("engine", ""))
            self.var_name.set(v.get("name", ""))
            self.var_year_f.set(str(v.get("year_from", 2000)))
            self.var_year_t.set(str(v.get("year_to", 9999)))
            self.var_can.set(v.get("can", "PT-CAN"))
            self.var_bitrate.set(str(v.get("bitrate", 500)))
        # BRL-Ausgabepfad vorbereiten wenn noch leer
        if lf and not self.var_brl_out.get():
            self.var_brl_out.set(str(lf.path.with_suffix(".brl")))
        self._refresh_preview()
        self._update_brl_src_label()

    def _on_select_sensor(self, _evt=None):
        lf = self._selected_file()
        sel = self.tree_sens.selection()
        if not lf or not sel:
            self._set_detail("")
            return
        try:
            idx = int(sel[0])
        except ValueError:
            return
        if not (0 <= idx < len(lf.sensors)):
            return
        s = lf.sensors[idx]
        if s.get("_empty"):
            self._set_detail(f"Slot {idx + 1}: leer")
            return
        lines = [f"Slot {s['_slot'] + 1}"]
        for k in FIELD_ORDER:
            lines.append(f"  {FIELD_LABELS_DE[k]:<18} {s.get(k, '')}")
        lines.append(f"  Protokoll (Klartext): {s.get('_proto_text', '')}")
        lines.append(f"  Typ (Klartext):       {s.get('_type_text', '')}")
        self._set_detail("\n".join(lines))

    def _set_detail(self, text: str):
        self.txt_detail.config(state="normal")
        self.txt_detail.delete("1.0", "end")
        self.txt_detail.insert("end", text)
        self.txt_detail.config(state="disabled")

    # -- Preview refresh ----------------------------------------------------

    def _refresh_preview(self):
        self.tree_sens.delete(*self.tree_sens.get_children())
        lf = self._selected_file()
        if not lf:
            self.lbl_vehicle.config(text="")
            self._set_detail("")
            return

        if lf.error:
            self.lbl_vehicle.config(text=f"Fehler: {lf.error}", fg=COL_RED)
            self._set_detail("")
            return

        # Fahrzeug-Info anzeigen (nur BRL)
        if lf.kind == "BRL" and lf.vehicle:
            v = lf.vehicle
            info = (f"BRL-Fahrzeug: {v.get('make','')} {v.get('model','')}  "
                    f"Motor: {v.get('engine','')}  -  {v.get('name','')}  "
                    f"({v.get('year_from','?')}-{v.get('year_to','?')})  "
                    f"{v.get('can','')} {v.get('bitrate','')} kbps")
            self.lbl_vehicle.config(text=info, fg=COL_AMBER)
        else:
            self.lbl_vehicle.config(text=(f"{lf.kind}-Header: {lf.header}"
                                          if lf.header else ""),
                                    fg=COL_GRAY)

        needle = self.var_filter.get().strip().lower()
        hide_empty = self.var_hide_empty.get()
        for i, s in enumerate(lf.sensors):
            if s.get("_empty") and hide_empty:
                continue
            if hide_empty and s.get("_leer"):
                continue
            if needle:
                hay = " ".join([
                    str(s.get("name", "")),
                    str(s.get("can_id", "")),
                    str(s.get("_proto_text", "")),
                    str(s.get("_type_text", "")),
                ]).lower()
                if needle not in hay:
                    continue
            if s.get("_empty"):
                self.tree_sens.insert(
                    "", "end", iid=str(i),
                    values=(s.get("_slot", i) + 1, "(leer)", "", "",
                            "", "", "", "", "", ""))
                continue
            self.tree_sens.insert(
                "", "end", iid=str(i),
                values=(
                    s.get("_slot", i) + 1,
                    s.get("name", ""),
                    s.get("_proto_text", ""),
                    s.get("can_id", ""),
                    s.get("len", ""),
                    s.get("scale", ""),
                    s.get("offset", ""),
                    s.get("min", ""),
                    s.get("max", ""),
                    s.get("_type_text", ""),
                ))

    # -- File-dialogs --------------------------------------------------------

    def _choose_excel_out(self):
        path = filedialog.asksaveasfilename(
            title="Excel speichern",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Alle Dateien", "*.*")],
            initialfile="can_sensoren.xlsx")
        if path:
            self.var_excel_out.set(path)

    def _choose_brl_out(self):
        initial = "profil.brl"
        lf = self._selected_file()
        if lf:
            initial = lf.path.stem + ".brl"
        path = filedialog.asksaveasfilename(
            title="BRL speichern",
            defaultextension=".brl",
            filetypes=[("BRL Profile", "*.brl"), ("Alle Dateien", "*.*")],
            initialfile=initial)
        if path:
            self.var_brl_out.set(path)

    # -- Excel export --------------------------------------------------------

    def _do_export_excel(self):
        loaded = [lf for lf in self._files if not lf.error]
        if not loaded:
            messagebox.showwarning("Keine Daten",
                                   "Bitte zuerst Dateien hinzufuegen.")
            return
        out = self.var_excel_out.get().strip()
        if not out:
            messagebox.showwarning("Kein Ziel", "Bitte Ausgabedatei angeben.")
            return
        self.btn_excel.config(state="disabled", text="Laeuft...")
        self.pb_excel["maximum"] = len(loaded)
        self.pb_excel["value"]   = 0
        threading.Thread(target=self._run_export_excel,
                         args=(loaded, Path(out)), daemon=True).start()

    def _run_export_excel(self, loaded, out_path):
        try:
            total = export_excel(loaded, out_path)
            self.after(0, self._excel_done, True,
                       f"{total} aktive Sensoren aus {len(loaded)} Dateien "
                       f"exportiert.", str(out_path))
        except Exception as e:
            self.after(0, self._excel_done, False, str(e), "")

    def _excel_done(self, ok, msg, path):
        self.btn_excel.config(state="normal", text="  EXCEL EXPORTIEREN")
        self.pb_excel["value"] = self.pb_excel["maximum"]
        if ok:
            self.lbl_excel.config(text=msg, fg=COL_GREEN)
            if messagebox.askyesno("Fertig",
                                   f"{msg}\n\nGespeichert: {path}\n\n"
                                   "Excel jetzt oeffnen?"):
                _open_file(path)
        else:
            self.lbl_excel.config(text=f"Fehler: {msg}", fg=COL_RED)
            messagebox.showerror("Fehler", f"Export fehlgeschlagen:\n\n{msg}")

    # -- BRL create ----------------------------------------------------------

    def _update_brl_src_label(self):
        """Zeigt im BRL-Tab an, aus welchen Sensoren gleich gebaut wird."""
        if not hasattr(self, "lbl_brl_src"):
            return  # BRL-Tab noch nicht gebaut
        mode = self.var_src_mode.get()
        if mode == "merge":
            sensors, dups = self._collect_merge_sensors()
            msg = f">>>  {len(sensors)} Sensoren aus {len([lf for lf in self._files if not lf.error])} Dateien"
            if dups > 0:
                msg += f"  ({dups} Duplikate uebersprungen)"
            self.lbl_brl_src.config(text=msg)
        else:
            lf = self._selected_file()
            if not lf or lf.error:
                self.lbl_brl_src.config(text=">>>  keine Quelldatei ausgewaehlt")
            else:
                self.lbl_brl_src.config(
                    text=f">>>  {lf.active_sensor_count()} Sensoren aus "
                         f"{lf.path.name} ({lf.kind})")

    def _collect_merge_sensors(self) -> tuple[list[dict], int]:
        """
        Vereint Sensoren aus ALLEN geladenen (nicht fehlerhaften) Dateien.
        Dedup anhand des Sensor-Namens (case-insensitiv, whitespace-getrimmt)
        - das erste Vorkommen gewinnt. So wird z.B. "RPM" nur einmal drin sein,
        egal ob es aus der OBD2- oder der OBD29-Datei kommt. Leere / "empty"
        Slots werden uebersprungen. Slots werden neu durchnummeriert 0..N-1.

        Rueckgabe: (sensor_list, anzahl_duplikate_verworfen)
        """
        seen: set[str] = set()
        merged: list[dict] = []
        dups = 0
        for lf in self._files:
            if lf.error:
                continue
            for s in lf.sensors:
                if s.get("_empty") or s.get("_leer"):
                    continue
                name = str(s.get("name", "")).strip().lower()
                if not name:
                    # Unbenannte Slots nicht dedupen (koennten echte,
                    # aber noch unbenannte Sensoren sein) — trotzdem mit
                    # eindeutigem Fallback-Key ablegen.
                    name = f"_unnamed_{len(merged)}"
                if name in seen:
                    dups += 1
                    continue
                seen.add(name)
                merged.append(dict(s))   # copy so we can rewrite _slot
        # Slots neu durchnummerieren
        for i, s in enumerate(merged):
            s["_slot"] = i
        return merged, dups

    def _do_create_brl(self):
        mode = self.var_src_mode.get()

        if mode == "merge":
            sensors, dups = self._collect_merge_sensors()
            if not sensors:
                messagebox.showwarning(
                    "Keine Sensoren",
                    "Keine aktiven Sensoren in den geladenen Dateien.")
                return
            src_desc = (f"{len(sensors)} Sensoren aus "
                        f"{len([lf for lf in self._files if not lf.error])} Dateien"
                        + (f" ({dups} Duplikate verworfen)" if dups else ""))
        else:
            lf = self._selected_file()
            if not lf:
                messagebox.showwarning(
                    "Keine Auswahl",
                    "Bitte in der Liste links eine Quelldatei "
                    "(.TRX/.TRI/.brl) auswaehlen oder auf 'Alle Dateien "
                    "zusammenfuehren' umschalten.")
                return
            if lf.error:
                messagebox.showerror(
                    "Datei-Fehler",
                    f"Datei konnte nicht geladen werden:\n{lf.error}")
                return
            sensors = lf.sensors
            src_desc = f"{lf.active_sensor_count()} Sensoren aus {lf.path.name}"

        if not self.var_engine.get().strip():
            messagebox.showwarning("Motor fehlt",
                                   "Bitte Motorcode eingeben (z.B. N47D20).")
            return
        if not self.var_name.get().strip():
            messagebox.showwarning("Name fehlt",
                                   "Bitte Bezeichnung eingeben "
                                   "(z.B. N47 2.0d 143PS).")
            return
        out = self.var_brl_out.get().strip()
        if not out:
            messagebox.showwarning("Kein Ziel", "Bitte Ausgabedatei angeben.")
            return
        vehicle = {
            "make":      self.var_make.get().strip() or "BMW",
            "model":     self.var_model.get().strip(),
            "engine":    self.var_engine.get().strip(),
            "name":      self.var_name.get().strip(),
            "year_from": int(self.var_year_f.get() or 0),
            "year_to":   int(self.var_year_t.get() or 9999),
            "can":       self.var_can.get(),
            "bitrate":   int(self.var_bitrate.get() or 500),
        }
        self.btn_brl_create.config(state="disabled", text="Erstelle...")
        threading.Thread(target=self._run_create_brl,
                         args=(sensors, vehicle, Path(out), src_desc),
                         daemon=True).start()

    def _run_create_brl(self, sensors, vehicle, out_path, src_desc):
        try:
            data = create_brl(vehicle, sensors)
            out_path.write_bytes(data)
            self.after(0, self._brl_done, True,
                       f"{src_desc}  ->  {out_path.name} "
                       f"({len(data) / 1024:.1f} KB)",
                       str(out_path))
        except Exception as e:
            self.after(0, self._brl_done, False, str(e), "")

    def _brl_done(self, ok, msg, path):
        self.btn_brl_create.config(state="normal", text="  BRL ERSTELLEN")
        if ok:
            self.lbl_brl.config(text=msg, fg=COL_GREEN)
            if messagebox.askyesno("Fertig",
                                   f"{msg}\n\nGespeichert: {path}\n\n"
                                   "Ordner oeffnen?"):
                _open_folder(path)
        else:
            self.lbl_brl.config(text=f"Fehler: {msg}", fg=COL_RED)
            messagebox.showerror("Fehler",
                                 f"Erstellen fehlgeschlagen:\n\n{msg}")


# ===========================================================================
# Externe Hilfsfunktionen
# ===========================================================================

def _open_file(path: str):
    # Immer fire-and-forget: subprocess.run() wartet auf Prozess-Ende und
    # blockiert damit den Tkinter-Mainloop, was die GUI einfriert. Popen
    # (ohne .wait()) loest das.
    import platform
    try:
        if platform.system() == "Windows":
            os.startfile(path)  # type: ignore[attr-defined]
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


def _open_folder(file_path: str):
    # Dasselbe Problem wie in _open_file: subprocess.run blockiert, bis
    # Explorer / open / xdg-open zurueckkehren. Auf Windows haengt sich
    # 'explorer /select,…' nach dem Export manchmal an die Tool-UI, GUI
    # friert ein. Popen laeuft asynchron, Tool bleibt reaktiv.
    import platform
    try:
        p = Path(file_path)
        if platform.system() == "Windows":
            subprocess.Popen(["explorer", "/select,", str(p)])
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", "-R", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p.parent)])
    except Exception:
        pass


# ===========================================================================
# Entry point
# ===========================================================================

if __name__ == "__main__":
    App().mainloop()
