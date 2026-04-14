#!/usr/bin/env python3
"""
CAN Checked TRX → Excel Konverter
===================================
Entschlüsselt eine oder mehrere .TRX Dateien und schreibt alle Sensoren
in eine Excel-Tabelle (.xlsx).

Jede TRX-Datei bekommt ein eigenes Tabellenblatt (Sheet).
Ein zusätzliches "Übersicht"-Blatt fasst alle aktiven Sensoren zusammen.

Verwendung:
  python3 trx_to_excel.py datei.TRX
  python3 trx_to_excel.py *.TRX -o ausgabe.xlsx
  python3 trx_to_excel.py --dir /pfad/zu/trx-dateien/ -o ausgabe.xlsx

Voraussetzungen:
  pip install pycryptodome openpyxl
"""

import sys
import os
import argparse
from pathlib import Path

# ── AES-256-ECB Schlüssel (aus MFD15 ESP32 Firmware, Offset 0x318C) ────────
TRX_KEY = bytes.fromhex(
    "3373357638792f423f4528482b4d6251655468576d5a7134743777397a244326"
)
# ASCII: 3s5v8y/B?E(H+MbQeThWmZq4t7w9z$C&

RECORD_SIZE = 192        # Bytes pro Sensor-Eintrag (verschlüsselt)
SENSOR_TEXT_SIZE = 96    # Bytes für den Text-Teil (null-terminiert)

# ── Spaltennamen (26 Felder, TRI Format) ────────────────────────────────────
FIELD_NAMES = [
    "Protokoll",        #  0  Header / Bus-Protokoll (hex)
    "CAN-ID",           #  1  CAN Bus Identifier (hex)
    "Format",           #  2  0=Big-endian, 1=Little-endian, 2=VEMS, 4=IEEE754
    "Startbyte",        #  3  Erstes Byte im CAN-Paket (0–7)
    "Länge",            #  4  Länge in Bytes (1, 2 oder 4)
    "Unsigned",         #  5  1=vorzeichenlos, 0=vorzeichenbehaftet
    "Shift",            #  6  Bit-Verschiebung nach rechts
    "CAN-Maske",        #  7  Bit-Maske (0 = keine Maske)
    "Dezimalstellen",   #  8  Anzahl Nachkommastellen
    "Sensorname",       #  9  Name (max. 15 Zeichen)
    "Skalierung",       # 10  initCalc  → Ergebnis = (Rohwert & Maske >> Shift) × Skalierung + Offset
    "Offset",           # 11  initOffset
    "Mapper-Typ",       # 12  0=linear, MAP=Kennlinie, NTC=Thermistor
    "Mapper-Info1",     # 13
    "Mapper-Info2",     # 14
    "Mapper-Info3",     # 15
    "Mapper-Info4",     # 16
    "AIN aktiv",        # 17  1=Analog-Eingang aktiv
    "Min-Warnung",      # 18  Unterer Schwellwert
    "Max-Warnung",      # 19  Oberer Schwellwert
    "Ref-Sensor",       # 20  255=kein Referenzsensor
    "Ref-Wert",         # 21
    "(ungenutzt1)",     # 22
    "Popup",            # 23  1=Popup bei Über-/Unterschreitung
    "(ungenutzt2)",     # 24
    "Sensor-Typ",       # 25  0=keine Einheit, 1=Druck, 2=Temp, 3=Geschw., 4=Lambda
]

SENSOR_TYPE_LABELS = {
    "0": "–",
    "1": "Druck",
    "2": "Temperatur",
    "3": "Geschwindigkeit",
    "4": "Lambda",
}

PROTOCOL_LABELS = {
    "0": "PT-CAN Broadcast",
    "0000": "PT-CAN Broadcast",
    "1": "BMW UDS Diagnose",
    "0001": "BMW UDS Diagnose",
    "7DF": "OBD2 Broadcast",
    "7df": "OBD2 Broadcast",
    "FFF": "MFD Analogeingang / Intern",
}


def _check_imports():
    missing = []
    try:
        from Crypto.Cipher import AES  # noqa: F401
    except ImportError:
        missing.append("pycryptodome")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print("Fehlende Pakete. Bitte installieren:", file=sys.stderr)
        print(f"  pip install {' '.join(missing)}", file=sys.stderr)
        sys.exit(1)


def decrypt_sensors(input_path: Path) -> tuple[str, list[dict]]:
    """
    Entschlüsselt eine TRX-Datei.

    Rückgabe
    --------
    (header_line, sensors)
      header_line : str  — Kopfzeile der TRX-Datei
      sensors     : list[dict] — Liste aller Sensor-Dicts (inkl. Leer-Sensoren)
    """
    from Crypto.Cipher import AES

    data = input_path.read_bytes()

    # Header-Ende suchen (\r\n für v1.0, \n für v1.1+)
    hdr_end = data.find(b"\r\n")
    if hdr_end != -1:
        hdr_end += 2
    else:
        hdr_end = data.find(b"\n")
        if hdr_end == -1:
            raise ValueError("Kein Zeilenende im Header gefunden")
        hdr_end += 1

    header_line = data[:hdr_end].decode("latin-1", errors="replace").strip()
    body = data[hdr_end:]

    if len(body) == 0:
        raise ValueError("Leerer Inhalt nach Header")
    if len(body) % 16 != 0:
        raise ValueError(
            f"Körperlänge {len(body)} nicht durch 16 teilbar — Datei möglicherweise beschädigt"
        )

    # AES-256-ECB entschlüsseln
    cipher = AES.new(TRX_KEY, AES.MODE_ECB)
    plaintext = cipher.decrypt(body)

    sensors = []
    num_records = len(plaintext) // RECORD_SIZE

    for i in range(num_records):
        text_block = plaintext[i * RECORD_SIZE : i * RECORD_SIZE + SENSOR_TEXT_SIZE]
        null_pos = text_block.find(b"\x00")
        if null_pos == -1:
            null_pos = SENSOR_TEXT_SIZE
        raw_text = text_block[:null_pos].decode("latin-1", errors="replace").rstrip(";")

        if not raw_text or ";" not in raw_text:
            sensors.append({"_slot": i, "_raw": "", "_empty": True})
            continue

        parts = raw_text.split(";")
        # Auf genau 26 Felder normalisieren
        while len(parts) < 26:
            parts.append("")

        sensor = {"_slot": i, "_raw": raw_text, "_empty": False}
        for j, name in enumerate(FIELD_NAMES):
            sensor[name] = parts[j] if j < len(parts) else ""

        # Hilfsspalten
        sensor["_leer"] = sensor.get("Sensorname", "").lower() == "empty"
        sensor["_protokoll_klartext"] = PROTOCOL_LABELS.get(
            sensor.get("Protokoll", ""), sensor.get("Protokoll", "")
        )
        sensor["_typ_klartext"] = SENSOR_TYPE_LABELS.get(
            sensor.get("Sensor-Typ", ""), sensor.get("Sensor-Typ", "")
        )
        sensors.append(sensor)

    return header_line, sensors


def _col_letter(n: int) -> str:
    """Spaltenindex (1-basiert) → Excel-Buchstabe(n)."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def build_excel(
    files_sensors: list[tuple[Path, str, list[dict]]],
    output_path: Path,
    only_active: bool = False,
):
    """Erstellt die Excel-Datei."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # Standard-Sheet entfernen

    # ── Farben ──────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dunkelblau
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    META_FILL  = PatternFill("solid", fgColor="D6E4F0")   # hellblau
    META_FONT  = Font(bold=True, size=10)
    EVEN_FILL  = PatternFill("solid", fgColor="F2F7FB")
    EMPTY_FONT = Font(color="AAAAAA", italic=True)
    THIN = Side(style="thin", color="CCCCCC")
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header_row(ws, row_num, num_cols):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def style_data_row(ws, row_num, num_cols, is_even, is_empty_sensor):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            if is_empty_sensor:
                cell.font = EMPTY_FONT
            if is_even:
                cell.fill = EVEN_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # ── Spalten für Sensor-Sheets ────────────────────────────────────────────
    SENSOR_COLS = (
        ["Slot", "Sensorname", "Protokoll", "Protokoll (Klartext)", "CAN-ID"]
        + FIELD_NAMES[2:]          # alle anderen Felder ab Index 2
        + ["Sensor-Typ (Klartext)"]
    )

    def write_sensor_sheet(ws, header_line, sensors, source_name, skip_empty=False):
        ws.freeze_panes = "A4"

        # Zeile 1: Metainfo
        ws["A1"] = f"Quelle: {source_name}"
        ws["A1"].font = META_FONT
        ws["A1"].fill = META_FILL
        ws.merge_cells(f"A1:{get_column_letter(len(SENSOR_COLS))}1")
        ws["A1"].alignment = Alignment(horizontal="left")

        ws["A2"] = f"TRX-Header: {header_line}"
        ws["A2"].font = Font(italic=True, size=9, color="555555")
        ws.merge_cells(f"A2:{get_column_letter(len(SENSOR_COLS))}2")

        # Zeile 3: Spaltenköpfe
        for ci, col in enumerate(SENSOR_COLS, start=1):
            ws.cell(row=3, column=ci, value=col)
        style_header_row(ws, 3, len(SENSOR_COLS))
        ws.row_dimensions[3].height = 30

        active_count = 0
        row = 4
        for s in sensors:
            if s.get("_empty"):
                continue
            is_leer = s.get("_leer", False)
            if skip_empty and is_leer:
                continue

            values = [
                s["_slot"] + 1,                        # Slot (1-basiert)
                s.get("Sensorname", ""),
                s.get("Protokoll", ""),
                s.get("_protokoll_klartext", ""),
                s.get("CAN-ID", ""),
            ]
            for fn in FIELD_NAMES[2:]:
                values.append(s.get(fn, ""))
            values.append(s.get("_typ_klartext", ""))

            for ci, val in enumerate(values, start=1):
                ws.cell(row=row, column=ci, value=val)

            style_data_row(ws, row, len(SENSOR_COLS), row % 2 == 0, is_leer)
            row += 1
            if not is_leer:
                active_count += 1

        # Spaltenbreiten anpassen
        col_widths = {
            1: 6,   # Slot
            2: 20,  # Sensorname
            3: 12,  # Protokoll
            4: 22,  # Protokoll Klartext
            5: 14,  # CAN-ID
            6: 9,   # Format
            7: 10,  # Startbyte
            8: 9,   # Länge
            9: 10,  # Unsigned
            10: 9,  # Shift
            11: 13, # CAN-Maske
            12: 14, # Dezimalstellen
            13: 18, # Skalierung
            14: 14, # Offset
        }
        for ci in range(1, len(SENSOR_COLS) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(ci, 14)

        return active_count

    # ── Übersicht-Sheet ──────────────────────────────────────────────────────
    overview_cols = ["Datei", "Slot"] + SENSOR_COLS[1:]  # Dateiname + alle Sensor-Cols

    ws_ov = wb.create_sheet("Übersicht")
    ws_ov.freeze_panes = "A3"
    ws_ov["A1"] = "Übersicht aller aktiven Sensoren"
    ws_ov["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws_ov.merge_cells(f"A1:{get_column_letter(len(overview_cols))}1")

    for ci, col in enumerate(overview_cols, start=1):
        ws_ov.cell(row=2, column=ci, value=col)
    style_header_row(ws_ov, 2, len(overview_cols))
    ws_ov.row_dimensions[2].height = 30

    ov_row = 3

    # ── Ein Sheet pro TRX-Datei ──────────────────────────────────────────────
    for trx_path, header_line, sensors in files_sensors:
        sheet_name = trx_path.stem[:31]   # Excel-Limit: 31 Zeichen
        # Doppelte Namen vermeiden
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + f"_{len(existing)}"

        ws = wb.create_sheet(sheet_name)
        write_sensor_sheet(ws, header_line, sensors, trx_path.name, skip_empty=False)

        # Aktive Sensoren in Übersicht eintragen
        for s in sensors:
            if s.get("_empty") or s.get("_leer"):
                continue
            values = [trx_path.name, s["_slot"] + 1, s.get("Sensorname", "")]
            values += [s.get("_protokoll_klartext", ""), s.get("CAN-ID", "")]
            for fn in FIELD_NAMES[2:]:
                values.append(s.get(fn, ""))
            values.append(s.get("_typ_klartext", ""))

            for ci, val in enumerate(values, start=1):
                ws_ov.cell(row=ov_row, column=ci, value=val)
            style_data_row(ws_ov, ov_row, len(overview_cols), ov_row % 2 == 0, False)
            ov_row += 1

    # Übersicht-Spaltenbreiten
    ov_widths = [20, 6, 20, 22, 14]
    for ci, w in enumerate(ov_widths, start=1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(len(ov_widths) + 1, len(overview_cols) + 1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = 14

    # Übersicht als erstes Sheet
    wb.move_sheet("Übersicht", offset=-len(wb.sheetnames) + 1)

    wb.save(output_path)
    print(f"\nExcel-Datei gespeichert: {output_path}")
    return ov_row - 3   # Anzahl aktiver Sensoren gesamt


def main():
    parser = argparse.ArgumentParser(
        description="CAN Checked .TRX → Excel Konverter",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        metavar="DATEI.TRX",
        help=".TRX Datei(en) zum Entschlüsseln",
    )
    parser.add_argument(
        "--dir",
        metavar="ORDNER",
        help="Alle .TRX Dateien in diesem Ordner verarbeiten",
    )
    parser.add_argument(
        "-o", "--output",
        metavar="AUSGABE.xlsx",
        default="trx_sensoren.xlsx",
        help="Name der Excel-Ausgabedatei (Standard: trx_sensoren.xlsx)",
    )
    args = parser.parse_args()

    _check_imports()

    # Eingabedateien sammeln
    input_files: list[Path] = []
    if args.dir:
        d = Path(args.dir)
        if not d.is_dir():
            print(f"FEHLER: Kein Ordner: {args.dir}", file=sys.stderr)
            sys.exit(1)
        input_files = sorted(d.glob("*.TRX")) + sorted(d.glob("*.trx"))
    for inp in args.inputs:
        p = Path(inp)
        if not p.exists():
            print(f"WARNUNG: Datei nicht gefunden: {inp}", file=sys.stderr)
            continue
        if p.is_dir():
            input_files += sorted(p.glob("*.TRX")) + sorted(p.glob("*.trx"))
        else:
            input_files.append(p)

    if not input_files:
        parser.print_help()
        sys.exit(0)

    # Entschlüsseln
    files_sensors = []
    for trx_path in input_files:
        try:
            header_line, sensors = decrypt_sensors(trx_path)
            active = sum(1 for s in sensors if not s.get("_empty") and not s.get("_leer"))
            print(f"  OK  {trx_path.name}  ({active} aktive Sensoren)")
            files_sensors.append((trx_path, header_line, sensors))
        except Exception as exc:
            print(f"FEHLER  {trx_path.name}: {exc}", file=sys.stderr)

    if not files_sensors:
        print("Keine Dateien erfolgreich entschlüsselt.", file=sys.stderr)
        sys.exit(1)

    # Excel erstellen
    out = Path(args.output)
    total = build_excel(files_sensors, out)
    print(f"Gesamt: {total} aktive Sensoren aus {len(files_sensors)} Datei(en).")


if __name__ == "__main__":
    main()
