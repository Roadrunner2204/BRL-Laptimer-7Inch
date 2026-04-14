#!/usr/bin/env python3
"""
CAN Checked TRX → Excel Konverter (GUI)
=========================================
Grafische Oberfläche zum Entschlüsseln von CAN Checked .TRX Dateien
und Export als Excel-Tabelle (.xlsx).

Voraussetzungen:
  pip install pycryptodome openpyxl

Starten:
  python3 trx_to_excel_gui.py
"""

import sys
import os
import json
import threading
import tempfile
import subprocess
import webbrowser
import urllib.request
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

# ── Version & Update ─────────────────────────────────────────────────────────
APP_VERSION = "1.0.0"
GITHUB_REPO = "roadrunner2204/bmw-data-display-v0.1"
SCRIPT_DIR  = Path(getattr(sys, '_MEIPASS', Path(__file__).resolve().parent))

# ── BMW Display Farbschema ───────────────────────────────────────────────────
COL_BG      = "#000000"
COL_PANEL   = "#1A1A1A"
COL_BORDER  = "#2C2C2E"
COL_BLUE    = "#1C69D4"
COL_BLUE_DK = "#1C3A6B"
COL_WHITE   = "#FFFFFF"
COL_GRAY    = "#AEAEB2"
COL_DARK    = "#636366"
COL_GREEN   = "#30D158"
COL_AMBER   = "#FF9F0A"
COL_RED     = "#FF453A"

# ── AES-256-ECB Schlüssel ────────────────────────────────────────────────────
TRX_KEY = bytes.fromhex(
    "3373357638792f423f4528482b4d6251655468576d5a7134743777397a244326"
)
RECORD_SIZE = 192
SENSOR_TEXT_SIZE = 96

FIELD_NAMES = [
    "Protokoll", "CAN-ID", "Format", "Startbyte", "Länge",
    "Unsigned", "Shift", "CAN-Maske", "Dezimalstellen", "Sensorname",
    "Skalierung", "Offset", "Mapper-Typ", "Mapper-Info1", "Mapper-Info2",
    "Mapper-Info3", "Mapper-Info4", "AIN aktiv", "Min-Warnung", "Max-Warnung",
    "Ref-Sensor", "Ref-Wert", "(ungenutzt1)", "Popup", "(ungenutzt2)", "Sensor-Typ",
]
SENSOR_TYPE_LABELS = {"0": "–", "1": "Druck", "2": "Temperatur", "3": "Geschwindigkeit", "4": "Lambda"}
PROTOCOL_LABELS = {
    "0": "PT-CAN Broadcast", "0000": "PT-CAN Broadcast",
    "1": "BMW UDS Diagnose", "0001": "BMW UDS Diagnose",
    "7DF": "OBD2 Broadcast", "7df": "OBD2 Broadcast",
    "FFF": "MFD Analogeingang / Intern",
}


# ── Entschlüsselung ──────────────────────────────────────────────────────────

def decrypt_sensors(input_path: Path):
    from Crypto.Cipher import AES
    data = input_path.read_bytes()
    hdr_end = data.find(b"\r\n")
    if hdr_end != -1:
        hdr_end += 2
    else:
        hdr_end = data.find(b"\n")
        if hdr_end == -1:
            raise ValueError("Kein Zeilenende im Header")
        hdr_end += 1

    header_line = data[:hdr_end].decode("latin-1", errors="replace").strip()
    body = data[hdr_end:]
    if not body:
        raise ValueError("Leerer Inhalt nach Header")
    if len(body) % 16 != 0:
        raise ValueError(f"Körperlänge {len(body)} nicht durch 16 teilbar")

    cipher = AES.new(TRX_KEY, AES.MODE_ECB)
    plaintext = cipher.decrypt(body)

    sensors = []
    for i in range(len(plaintext) // RECORD_SIZE):
        block = plaintext[i * RECORD_SIZE: i * RECORD_SIZE + SENSOR_TEXT_SIZE]
        null = block.find(b"\x00")
        text = block[:null if null != -1 else SENSOR_TEXT_SIZE].decode("latin-1", errors="replace").rstrip(";")
        if not text or ";" not in text:
            sensors.append({"_slot": i, "_empty": True})
            continue
        parts = text.split(";")
        while len(parts) < 26:
            parts.append("")
        s = {"_slot": i, "_empty": False, "_raw": text}
        for j, n in enumerate(FIELD_NAMES):
            s[n] = parts[j] if j < len(parts) else ""
        s["_leer"] = s.get("Sensorname", "").lower() == "empty"
        s["_protokoll_klartext"] = PROTOCOL_LABELS.get(s.get("Protokoll", ""), s.get("Protokoll", ""))
        s["_typ_klartext"] = SENSOR_TYPE_LABELS.get(s.get("Sensor-Typ", ""), s.get("Sensor-Typ", ""))
        sensors.append(s)

    return header_line, sensors


# ── Excel-Export ─────────────────────────────────────────────────────────────

def build_excel(files_sensors, output_path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    META_FILL = PatternFill("solid", fgColor="D6E4F0")
    EVEN_FILL = PatternFill("solid", fgColor="F2F7FB")
    THIN = Side(style="thin", color="CCCCCC")
    TB = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hdr(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = TB
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30

    def data_row(ws, row, ncols, even, grey):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            if grey: cell.font = Font(color="AAAAAA", italic=True)
            if even: cell.fill = EVEN_FILL
            cell.border = TB
            cell.alignment = Alignment(vertical="center")

    SCOLS = (["Slot", "Sensorname", "Protokoll", "Protokoll (Klartext)", "CAN-ID"]
             + FIELD_NAMES[2:] + ["Sensor-Typ (Klartext)"])
    OCOLS = ["Datei"] + SCOLS

    ws_ov = wb.create_sheet("Übersicht")
    ws_ov.freeze_panes = "A3"
    ws_ov["A1"] = "Übersicht aller aktiven Sensoren"
    ws_ov["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws_ov.merge_cells(f"A1:{get_column_letter(len(OCOLS))}1")
    for ci, col in enumerate(OCOLS, 1):
        ws_ov.cell(row=2, column=ci, value=col)
    hdr(ws_ov, 2, len(OCOLS))
    ov_row = 3

    for trx_path, header_line, sensors in files_sensors:
        sheet_name = trx_path.stem[:31]
        existing = [s.title for s in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:27] + f"_{len(existing)}"

        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A4"
        ws["A1"] = f"Quelle: {trx_path.name}"
        ws["A1"].font = Font(bold=True, size=10); ws["A1"].fill = META_FILL
        ws.merge_cells(f"A1:{get_column_letter(len(SCOLS))}1")
        ws["A2"] = f"TRX-Header: {header_line}"
        ws["A2"].font = Font(italic=True, size=9, color="555555")
        ws.merge_cells(f"A2:{get_column_letter(len(SCOLS))}2")
        for ci, col in enumerate(SCOLS, 1):
            ws.cell(row=3, column=ci, value=col)
        hdr(ws, 3, len(SCOLS))

        row = 4
        for s in sensors:
            if s.get("_empty"): continue
            is_leer = s.get("_leer", False)
            vals = [s["_slot"]+1, s.get("Sensorname",""), s.get("Protokoll",""),
                    s.get("_protokoll_klartext",""), s.get("CAN-ID","")]
            for fn in FIELD_NAMES[2:]: vals.append(s.get(fn,""))
            vals.append(s.get("_typ_klartext",""))
            for ci, v in enumerate(vals, 1): ws.cell(row=row, column=ci, value=v)
            data_row(ws, row, len(SCOLS), row % 2 == 0, is_leer)
            if not is_leer:
                ov_vals = [trx_path.name] + vals
                for ci, v in enumerate(ov_vals, 1): ws_ov.cell(row=ov_row, column=ci, value=v)
                data_row(ws_ov, ov_row, len(OCOLS), ov_row % 2 == 0, False)
                ov_row += 1
            row += 1

        col_w = {1:6, 2:20, 3:12, 4:22, 5:14}
        for ci in range(1, len(SCOLS)+1):
            ws.column_dimensions[get_column_letter(ci)].width = col_w.get(ci, 14)

    for ci, w in enumerate([20, 6, 20, 22, 14], 1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(6, len(OCOLS)+1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = 14

    wb.move_sheet("Übersicht", offset=-len(wb.sheetnames)+1)
    wb.save(output_path)
    return ov_row - 3


# ── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"BMW Data Display  —  TRX Excel Konverter  v{APP_VERSION}")
        self.resizable(True, True)
        self.minsize(700, 540)
        self.configure(bg=COL_BG)
        self._files: list[Path] = []
        self._setup_style()
        self._build_ui()
        self._check_deps()
        self.after(2000, lambda: self._check_updates(silent=True))

    # ── ttk Style ────────────────────────────────────────────────────────────
    def _setup_style(self):
        s = ttk.Style(self)
        try:
            s.theme_use('clam')
        except Exception:
            pass
        s.configure("BMW.Vertical.TScrollbar",
                    background=COL_PANEL, troughcolor=COL_BG,
                    bordercolor=COL_BORDER, arrowcolor=COL_DARK, relief="flat")
        s.configure("BMW.Horizontal.TProgressbar",
                    background=COL_BLUE, troughcolor=COL_PANEL,
                    bordercolor=COL_BORDER, lightcolor=COL_BLUE,
                    darkcolor=COL_BLUE, relief="flat")

    # ── UI aufbauen ──────────────────────────────────────────────────────────
    def _build_ui(self):

        # ── Kopfzeile ────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=COL_BG)
        hdr.pack(fill="x")

        tk.Frame(hdr, bg=COL_BLUE, width=6).pack(side="left", fill="y")

        logo_frame = tk.Frame(hdr, bg=COL_BG, padx=14, pady=10)
        logo_frame.pack(side="left", fill="y")
        self._build_logo(logo_frame)

        title_f = tk.Frame(hdr, bg=COL_BG, padx=8, pady=10)
        title_f.pack(side="left", fill="both", expand=True)
        tk.Label(title_f, text="BMW DATA DISPLAY",
                 bg=COL_BG, fg=COL_WHITE,
                 font=("Helvetica", 18, "bold")).pack(anchor="w")
        tk.Label(title_f,
                 text="TRX Excel Konverter  –  Sensorkonfigurationen entschlüsseln & exportieren",
                 bg=COL_BG, fg=COL_GRAY, font=("Helvetica", 9)).pack(anchor="w")

        tk.Label(hdr, text=f"v{APP_VERSION}",
                 bg=COL_BG, fg=COL_BLUE,
                 font=("Helvetica", 11, "bold"),
                 padx=18).pack(side="right", fill="y")

        tk.Frame(self, bg=COL_BORDER, height=1).pack(fill="x")

        # ── Hauptinhalt ──────────────────────────────────────────────────────
        main = tk.Frame(self, bg=COL_BG)
        main.pack(fill="both", expand=True, padx=14, pady=10)

        # Panel: TRX-Dateien
        fp = tk.LabelFrame(main, text="  TRX-DATEIEN  ",
                           bg=COL_PANEL, fg=COL_BLUE,
                           font=("Helvetica", 9, "bold"),
                           bd=1, relief="flat",
                           highlightbackground=COL_BORDER,
                           highlightcolor=COL_BLUE,
                           highlightthickness=1)
        fp.pack(fill="both", expand=True, pady=(0, 10))

        btn_row = tk.Frame(fp, bg=COL_PANEL)
        btn_row.pack(fill="x", padx=10, pady=(10, 6))
        tk.Button(btn_row, text="  Dateien hinzufügen", command=self._add_files,
                  bg=COL_BLUE, fg=COL_WHITE, activebackground=COL_BLUE_DK,
                  activeforeground=COL_WHITE, font=("Helvetica", 10, "bold"),
                  relief="flat", padx=12, pady=7, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(btn_row, text="  Ordner hinzufügen", command=self._add_folder,
                  bg=COL_BLUE_DK, fg=COL_WHITE, activebackground="#0F2040",
                  activeforeground=COL_WHITE, font=("Helvetica", 10),
                  relief="flat", padx=12, pady=7, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(btn_row, text="Liste leeren", command=self._clear_files,
                  bg="#3A1010", fg=COL_WHITE, activebackground="#5A1515",
                  activeforeground=COL_WHITE, font=("Helvetica", 10),
                  relief="flat", padx=12, pady=7, cursor="hand2").pack(side="right")

        list_frame = tk.Frame(fp, bg=COL_PANEL)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        self.listbox = tk.Listbox(
            list_frame, selectmode="extended",
            font=("Courier", 10),
            bg="#0D0D0D", fg=COL_GRAY,
            selectbackground=COL_BLUE, selectforeground=COL_WHITE,
            relief="flat", bd=0,
            highlightthickness=1,
            highlightbackground=COL_BORDER,
            highlightcolor=COL_BLUE,
            activestyle="none")
        sb = ttk.Scrollbar(list_frame, orient="vertical",
                           command=self.listbox.yview,
                           style="BMW.Vertical.TScrollbar")
        self.listbox.configure(yscrollcommand=sb.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.lbl_count = tk.Label(fp, text="0 Datei(en) ausgewählt",
                                   bg=COL_PANEL, fg=COL_DARK, font=("Helvetica", 9))
        self.lbl_count.pack(anchor="w", padx=12, pady=(0, 8))

        # Panel: Ausgabe
        op = tk.LabelFrame(main, text="  EXCEL-AUSGABE  ",
                           bg=COL_PANEL, fg=COL_BLUE,
                           font=("Helvetica", 9, "bold"),
                           bd=1, relief="flat",
                           highlightbackground=COL_BORDER,
                           highlightcolor=COL_BLUE,
                           highlightthickness=1)
        op.pack(fill="x", pady=(0, 10))

        out_row = tk.Frame(op, bg=COL_PANEL)
        out_row.pack(fill="x", padx=10, pady=10)
        self.var_out = tk.StringVar(value=str(Path.home() / "trx_sensoren.xlsx"))
        tk.Entry(out_row, textvariable=self.var_out, font=("Helvetica", 10),
                 bg="#0D0D0D", fg=COL_GRAY, insertbackground=COL_BLUE,
                 relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=COL_BORDER,
                 highlightcolor=COL_BLUE).pack(side="left", fill="x", expand=True,
                                               ipady=6, padx=(0, 8))
        tk.Button(out_row, text=" … ", command=self._choose_output,
                  bg=COL_PANEL, fg=COL_GRAY, activebackground=COL_BORDER,
                  relief="flat", padx=8, cursor="hand2").pack(side="right")

        # ── Aktionsleiste ─────────────────────────────────────────────────────
        action = tk.Frame(self, bg=COL_BG, padx=14, pady=8)
        action.pack(fill="x")

        self.btn_start = tk.Button(action, text="  EXPORTIEREN",
                                    command=self._start_export,
                                    bg=COL_GREEN, fg=COL_WHITE,
                                    activebackground="#1E8A3C", activeforeground=COL_WHITE,
                                    font=("Helvetica", 11, "bold"),
                                    relief="flat", padx=20, pady=10, cursor="hand2")
        self.btn_start.pack(side="left")

        self.lbl_status = tk.Label(action, text="",
                                    bg=COL_BG, fg=COL_GRAY, font=("Helvetica", 10))
        self.lbl_status.pack(side="left", padx=16)

        tk.Button(action, text="Updates prüfen", command=self._check_updates,
                  bg=COL_PANEL, fg=COL_GRAY, activebackground=COL_BORDER,
                  font=("Helvetica", 9), relief="flat", padx=10, pady=10,
                  cursor="hand2").pack(side="right")

        self.progress = ttk.Progressbar(self, mode="determinate",
                                         style="BMW.Horizontal.TProgressbar")
        self.progress.pack(fill="x", padx=14, pady=(0, 6))

        # ── Fußzeile ─────────────────────────────────────────────────────────
        tk.Frame(self, bg=COL_BORDER, height=1).pack(fill="x")
        tk.Label(self, text=f"github.com/{GITHUB_REPO}",
                 bg=COL_BG, fg=COL_DARK, font=("Helvetica", 8)).pack(pady=4)

    # ── Logo ─────────────────────────────────────────────────────────────────
    def _build_logo(self, parent):
        logo_path = SCRIPT_DIR / "logo.png"
        if logo_path.exists():
            try:
                img = tk.PhotoImage(file=str(logo_path))
                w, h = img.width(), img.height()
                factor = max(1, max(w, h) // 64)
                if factor > 1:
                    img = img.subsample(factor, factor)
                self._logo_img = img
                tk.Label(parent, image=img, bg=COL_BG).pack()
                return
            except Exception:
                pass
        # Fallback: BMW Roundel
        size = 58
        r = size // 2
        cv = tk.Canvas(parent, width=size, height=size,
                       bg=COL_BG, highlightthickness=0)
        cv.pack()
        m = 2
        cv.create_oval(m, m, size - m, size - m,
                       fill="#0D1B2A", outline=COL_BLUE, width=2)
        ri = r - 6
        cx = cy = r
        # oben-links=blau, oben-rechts=weiß, unten-rechts=blau, unten-links=weiß
        cv.create_arc(cx-ri, cy-ri, cx+ri, cy+ri,
                      start=90, extent=90, fill=COL_BLUE, outline="", style="pieslice")
        cv.create_arc(cx-ri, cy-ri, cx+ri, cy+ri,
                      start=0, extent=90, fill=COL_WHITE, outline="", style="pieslice")
        cv.create_arc(cx-ri, cy-ri, cx+ri, cy+ri,
                      start=270, extent=90, fill=COL_BLUE, outline="", style="pieslice")
        cv.create_arc(cx-ri, cy-ri, cx+ri, cy+ri,
                      start=180, extent=90, fill=COL_WHITE, outline="", style="pieslice")
        cv.create_line(cx, cy - ri, cx, cy + ri, fill="#0D1B2A", width=2)
        cv.create_line(cx - ri, cy, cx + ri, cy, fill="#0D1B2A", width=2)
        tk.Label(parent, text="BMW", bg=COL_BG, fg=COL_BLUE,
                 font=("Helvetica", 8, "bold")).pack()

    # ── Abhängigkeiten prüfen ─────────────────────────────────────────────────
    def _check_deps(self):
        missing = []
        try:
            from Crypto.Cipher import AES  # noqa
        except ImportError:
            missing.append("pycryptodome")
        try:
            import openpyxl  # noqa
        except ImportError:
            missing.append("openpyxl")
        if missing:
            messagebox.showerror(
                "Fehlende Pakete",
                f"Bitte installiere:\n\n    pip install {' '.join(missing)}\n\nDann neu starten."
            )
            self.destroy()

    # ── Datei-Aktionen ────────────────────────────────────────────────────────
    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="TRX-Dateien auswählen",
            filetypes=[("TRX Dateien", "*.TRX *.trx"), ("Alle Dateien", "*.*")]
        )
        for p in paths:
            self._add_path(Path(p))
        self._refresh_list()

    def _add_folder(self):
        d = filedialog.askdirectory(title="Ordner mit TRX-Dateien auswählen")
        if d:
            found = sorted(Path(d).glob("*.TRX")) + sorted(Path(d).glob("*.trx"))
            if not found:
                messagebox.showinfo("Keine Dateien", f"Keine .TRX Dateien in:\n{d}")
                return
            for p in found:
                self._add_path(p)
            self._refresh_list()

    def _add_path(self, p: Path):
        if p not in self._files:
            self._files.append(p)

    def _clear_files(self):
        self._files.clear()
        self._refresh_list()

    def _refresh_list(self):
        self.listbox.delete(0, "end")
        for p in self._files:
            self.listbox.insert("end", f"  {p.name}   ({p.parent})")
        self.lbl_count.config(text=f"{len(self._files)} Datei(en) ausgewählt")

    def _choose_output(self):
        path = filedialog.asksaveasfilename(
            title="Excel-Datei speichern",
            defaultextension=".xlsx",
            filetypes=[("Excel Dateien", "*.xlsx"), ("Alle Dateien", "*.*")],
            initialfile="trx_sensoren.xlsx",
        )
        if path:
            self.var_out.set(path)

    # ── Export ────────────────────────────────────────────────────────────────
    def _start_export(self):
        if not self._files:
            messagebox.showwarning("Keine Dateien", "Bitte zuerst .TRX Dateien hinzufügen.")
            return
        out = self.var_out.get().strip()
        if not out:
            messagebox.showwarning("Kein Ziel", "Bitte eine Ausgabedatei angeben.")
            return
        self.btn_start.config(state="disabled", text="Läuft…")
        self.progress["maximum"] = len(self._files)
        self.progress["value"] = 0
        threading.Thread(target=self._run_export,
                         args=(list(self._files), Path(out)), daemon=True).start()

    def _run_export(self, files, out_path):
        results, errors = [], []
        for i, f in enumerate(files):
            try:
                header, sensors = decrypt_sensors(f)
                active = sum(1 for s in sensors if not s.get("_empty") and not s.get("_leer"))
                results.append((f, header, sensors))
                self.after(0, self._update_status,
                           f"Entschlüsselt: {f.name}  ({active} Sensoren)", i + 1, COL_GRAY)
            except Exception as e:
                errors.append(f"{f.name}: {e}")
                self.after(0, self._update_status, f"Fehler: {f.name}", i + 1, COL_RED)
        if not results:
            self.after(0, self._export_done, False, errors, 0)
            return
        try:
            total = build_excel(results, out_path)
            self.after(0, self._export_done, True, errors, total, out_path, len(results))
        except Exception as e:
            self.after(0, self._export_done, False, [str(e)], 0)

    def _update_status(self, msg, progress, color=None):
        self.lbl_status.config(text=msg, fg=color or COL_GRAY)
        self.progress["value"] = progress

    def _export_done(self, success, errors, total, out_path=None, n_files=None):
        self.btn_start.config(state="normal", text="  EXPORTIEREN")
        if success:
            msg = f"{total} aktive Sensoren aus {n_files} Datei(en) exportiert."
            self.lbl_status.config(text=msg, fg=COL_GREEN)
            err_text = ("\n\nFehler bei:\n" + "\n".join(errors)) if errors else ""
            if messagebox.askyesno(
                "Export abgeschlossen",
                f"{msg}\n\nGespeichert in:\n{out_path}{err_text}\n\nExcel jetzt öffnen?"
            ):
                self._open_file(out_path)
        else:
            self.lbl_status.config(text="Export fehlgeschlagen", fg=COL_RED)
            messagebox.showerror("Fehler", "Export fehlgeschlagen:\n\n" + "\n".join(errors))

    @staticmethod
    def _open_file(path):
        import platform
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)])
            else:
                subprocess.run(["xdg-open", str(path)])
        except Exception:
            pass

    # ── Update-Check ─────────────────────────────────────────────────────────
    def _check_updates(self, silent=False):
        threading.Thread(target=self._fetch_update, args=(silent,), daemon=True).start()

    def _fetch_update(self, silent):
        url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "TRX-Excel-Konverter"})
            with urllib.request.urlopen(req, timeout=8) as r:
                data = json.loads(r.read())
            latest = data.get("tag_name", "").lstrip("v")
            if not latest:
                return
            installer_url = None
            for asset in data.get("assets", []):
                name = asset.get("name", "")
                if name.lower().endswith(".exe") and "setup" in name.lower():
                    installer_url = asset.get("browser_download_url")
                    break
            changelog = data.get("body", "")[:600]
            def ver(s):
                try: return tuple(int(x) for x in s.split("."))
                except ValueError: return (0,)
            if ver(latest) > ver(APP_VERSION):
                self.after(0, self._show_update_dialog, latest, installer_url, changelog)
            elif not silent:
                self.after(0, messagebox.showinfo, "Kein Update",
                           f"Du hast die neueste Version (v{APP_VERSION}).")
        except Exception as e:
            if not silent:
                self.after(0, messagebox.showwarning, "Update-Prüfung",
                           f"GitHub nicht erreichbar:\n{e}")

    def _show_update_dialog(self, latest_version, installer_url, changelog):
        dlg = tk.Toplevel(self)
        dlg.title(f"Update verfügbar  –  v{latest_version}")
        dlg.configure(bg=COL_BG)
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Frame(dlg, bg=COL_BLUE, width=5).pack(side="left", fill="y")

        content = tk.Frame(dlg, bg=COL_BG, padx=24, pady=20)
        content.pack(fill="both", expand=True)

        tk.Label(content, text="UPDATE VERFÜGBAR",
                 bg=COL_BG, fg=COL_BLUE,
                 font=("Helvetica", 13, "bold")).pack(anchor="w")
        tk.Label(content,
                 text=f"Version {latest_version} verfügbar  (aktuell: v{APP_VERSION})",
                 bg=COL_BG, fg=COL_GRAY,
                 font=("Helvetica", 10)).pack(anchor="w", pady=(2, 14))

        if changelog:
            cl_f = tk.Frame(content, bg=COL_PANEL,
                            highlightbackground=COL_BORDER, highlightthickness=1)
            cl_f.pack(fill="x", pady=(0, 14))
            cl_txt = tk.Text(cl_f, height=6, wrap="word",
                             bg=COL_PANEL, fg=COL_GRAY,
                             font=("Helvetica", 9), relief="flat", bd=0,
                             padx=10, pady=8)
            cl_txt.insert("end", changelog)
            cl_txt.config(state="disabled")
            cl_txt.pack(fill="x")

        dl_info = tk.StringVar(value="")
        lbl_dl  = tk.Label(content, textvariable=dl_info,
                           bg=COL_BG, fg=COL_AMBER, font=("Helvetica", 9))
        dl_bar  = ttk.Progressbar(content, mode="determinate",
                                   style="BMW.Horizontal.TProgressbar")

        btn_row = tk.Frame(content, bg=COL_BG)
        btn_row.pack(fill="x", pady=(4, 0))

        def do_later():
            dlg.destroy()

        def do_install():
            if not installer_url:
                messagebox.showwarning(
                    "Kein Installer",
                    "Kein Windows-Installer in diesem Release gefunden.\n\n"
                    "Bitte auf GitHub manuell herunterladen.",
                    parent=dlg)
                webbrowser.open(f"https://github.com/{GITHUB_REPO}/releases/latest")
                dlg.destroy()
                return
            btn_inst.config(state="disabled", text="Wird heruntergeladen…")
            btn_late.config(state="disabled")
            lbl_dl.pack(anchor="w", pady=(8, 2))
            dl_bar.pack(fill="x", pady=(0, 8))

            def reporthook(pct, mb, mb_total):
                dl_info.set(f"Herunterladen…  {mb:.1f} / {mb_total:.1f} MB  ({pct}%)")
                dl_bar["value"] = pct

            def on_done(tmp_path):
                dl_info.set("Download abgeschlossen – Installer wird gestartet…")
                dlg.after(800, lambda: self._run_installer(tmp_path, dlg))

            def on_error(msg):
                btn_inst.config(state="normal", text="  Update installieren")
                btn_late.config(state="normal")
                dl_info.set(f"Fehler: {msg}")
                lbl_dl.config(fg=COL_RED)

            threading.Thread(
                target=self._do_download,
                args=(installer_url, latest_version, reporthook, on_done, on_error),
                daemon=True).start()

        btn_inst = tk.Button(btn_row, text="  Update installieren",
                             command=do_install,
                             bg=COL_BLUE, fg=COL_WHITE,
                             activebackground=COL_BLUE_DK, activeforeground=COL_WHITE,
                             font=("Helvetica", 11, "bold"),
                             relief="flat", padx=16, pady=10, cursor="hand2")
        btn_inst.pack(side="left")

        btn_late = tk.Button(btn_row, text="Später",
                             command=do_later,
                             bg=COL_PANEL, fg=COL_GRAY,
                             activebackground=COL_BORDER, activeforeground=COL_WHITE,
                             font=("Helvetica", 10),
                             relief="flat", padx=14, pady=10, cursor="hand2")
        btn_late.pack(side="right")

        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - dlg.winfo_reqwidth())  // 2
        y = self.winfo_y() + (self.winfo_height() - dlg.winfo_reqheight()) // 2
        dlg.geometry(f"+{x}+{y}")

    # ── Download & Install ────────────────────────────────────────────────────
    @staticmethod
    def _do_download(url, version, reporthook, on_done, on_error):
        try:
            tmp_dir  = Path(tempfile.mkdtemp())
            tmp_path = tmp_dir / f"TRX-Excel-Konverter-Setup-v{version}.exe"
            req = urllib.request.Request(
                url, headers={"User-Agent": "TRX-Excel-Konverter"})
            with urllib.request.urlopen(req, timeout=60) as resp:
                total = int(resp.headers.get("Content-Length", 0))
                done  = 0
                with open(tmp_path, "wb") as f:
                    while True:
                        chunk = resp.read(65536)
                        if not chunk:
                            break
                        f.write(chunk)
                        done += len(chunk)
                        if total:
                            reporthook(min(100, done * 100 // total),
                                       done / 1_000_000, total / 1_000_000)
            on_done(tmp_path)
        except Exception as e:
            on_error(str(e))

    def _run_installer(self, tmp_path: Path, dialog=None):
        if dialog:
            try:
                dialog.destroy()
            except Exception:
                pass
        try:
            subprocess.Popen([
                str(tmp_path),
                "/SILENT",
                "/CLOSEAPPLICATIONS",
                "/RESTARTAPPLICATIONS",
            ])
        except Exception as e:
            messagebox.showerror(
                "Installer-Fehler",
                f"Installer konnte nicht gestartet werden:\n{e}\n\nDatei: {tmp_path}")
            return
        self.after(500, self.destroy)


# ── Start ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
